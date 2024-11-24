from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime

# Setup WebDriver
options = webdriver.ChromeOptions()
options.headless = False  # Set to True if you don't want the browser to open
driver = webdriver.Chrome(options=options)

try:
    # Navigate to Google
    driver.get("https://www.google.com")

    # Handle consent dialog (if applicable)
    try:
        consent_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#L2AGLb, button[aria-label='Accept all']"))
        )
        consent_button.click()
        print("Consent dialog accepted.")
    except Exception as e:
        print(f"No consent dialog found: {e}")

    # Open Excel file
    workbook_path = "D:/test/data.xlsx"
    try:
        workbook = load_workbook(workbook_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found at {workbook_path}. Please check the path.")

    # Get the sheet for the current day
    today = datetime.now().strftime('%A')
    if today in workbook.sheetnames:
        sheet = workbook[today]
    else:
        raise KeyError(f"Worksheet '{today}' does not exist in the Excel file. Available sheets: {workbook.sheetnames}")

    # Iterate through keywords
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
        keyword = row[0]
        if not keyword:
            print(f"Skipping empty keyword in row {row_idx}.")
            continue

        # Perform Google search
        driver.get("https://www.google.com")
        try:
            # Locate the search box, type the keyword, and wait for suggestions
            search_box = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, "q"))
            )
            search_box.clear()
            search_box.send_keys(keyword)

            # Wait for suggestions to load
            suggestions = WebDriverWait(driver, 10).until(
                EC.visibility_of_all_elements_located((By.CSS_SELECTOR, ".erkvQe .sbct"))
            )

            # Extract and process suggestions
            options = [s.text.strip() for s in suggestions if s.text.strip()]
            if not options:
                print(f"No suggestions found for '{keyword}'. Skipping.")
                continue

            # Find the longest and shortest suggestions
            longest = max(options, key=len)
            shortest = min(options, key=len)

            # Write results back to Excel
            sheet.cell(row=row_idx, column=2, value=longest)
            sheet.cell(row=row_idx, column=3, value=shortest)
            print(f"Suggestions for '{keyword}': Longest: {longest}, Shortest: {shortest}")

        except Exception as e:
            print(f"Error extracting suggestions for '{keyword}': {e}")
            continue

    # Save the updated Excel file
    updated_file_path = "D:/test/updated_data.xlsx"
    workbook.save(updated_file_path)
    print(f"Updated Excel file saved at {updated_file_path}.")

finally:
    # Quit the WebDriver
    driver.quit()
